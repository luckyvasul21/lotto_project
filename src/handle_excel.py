from multiprocessing import context

import openpyxl
import os

from src.common_methods import get_draw_date
from src.get_lotto_results import lotto_go

bookname = 'Sample.xlsx'
sheetname = 'Sample'

def load_workbook(bookname):
    filename = os.path.join(os.getcwd()+'\\resources\\', bookname)
    return openpyxl.load_workbook(filename)

# def load_sheet(bookname, sheetname):
#     book = load_workbook(bookname)
#     return book[sheetname]

def get_max_row(sheet):
    return sheet.max_row

def get_max_column(sheet):
    return sheet.max_column

def get_sheetnames(bookname):
    print(load_workbook(bookname).sheetnames)

def read_excel(bookname, sheetname):
    # Get a sheet to read
    sheet = load_workbook(bookname)[sheetname]
    # Reading each cell in excel
    for i in range(1, get_max_row(sheet)+1):
        for j in range(1, get_max_column(sheet)+1):
            print(sheet.cell(row=i, column=j).value)

def write_excel(context, bookname, sheetname):
    lotto_data_this_week = lotto_go(context)

    if 'euromillion' in context.browser.current_url:
        lotto_data_this_week.append('x')
        lotto_data_this_week.append('x')


# add draw date into array
    lotto_data_this_week.append(get_draw_date(context))
    # print(lotto_data_this_week)


    book = load_workbook(bookname)
    sheet = book[sheetname]

    # print(get_max_row(sheet))
    get_last_row = get_max_row(sheet)+1
    # print(get_last_row)

    for k in range(0, len(lotto_data_this_week)):
        # print(lotto_data_this_week[k])
        sheet.cell(row=get_last_row, column=k+1).value = lotto_data_this_week[k]

        book.save(os.path.join(os.getcwd()+'\\resources\\', bookname))
        book.close()

if __name__ == '__main__':
    # read_excel('Sample.xlsx', 'Sample')
    write_excel(context, 'Sample.xlsx', 'Sample')